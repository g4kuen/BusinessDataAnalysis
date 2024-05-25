import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as Img

df = pd.read_csv('preprocessed_data.csv', delimiter=',', header=None,
                 names=['updated', 'high', 'low', 'open', 'close', 'volume', 'average', 'U/R'])

def fifth():
    df_close_higher = df[df['close'].astype(float) > df['open'].astype(float)]
    df_open_higher = df[df['open'].astype(float) > df['close'].astype(float)]


    wb = openpyxl.Workbook()


    ws_close = wb.active
    ws_close.title = "Close higher"
    for r_idx, row in enumerate(dataframe_to_rows(df_close_higher, index=False), 1):
        for c_idx, value in enumerate(row, 1):
            ws_close.cell(row=r_idx, column=c_idx, value=value)

    ws_open = wb.create_sheet("Open higher")
    for r_idx, row in enumerate(dataframe_to_rows(df_open_higher, index=False), 1):
        for c_idx, value in enumerate(row, 1):
            ws_open.cell(row=r_idx, column=c_idx, value=value)

    fig_close, ax_close = plt.subplots()
    ax_close.hist(df_close_higher['volume'], bins=50)
    ax_close.set_xlabel('Объем торгов')
    ax_close.set_ylabel('Частота')
    ax_close.set_title('Гистограмма объемов торгов для дат, у которых точка закрытия выше точки открытия')
    plt.savefig('histogram_close.png')

    fig_open, ax_open = plt.subplots()
    ax_open.hist(df_open_higher['volume'], bins=50)
    ax_open.set_xlabel('Объем торгов')
    ax_open.set_ylabel('Частота')
    ax_open.set_title('Гистограмма объемов торгов для дат, у которых точка закрытия ниже точки открытия')
    plt.savefig('histogram_open.png')

    img_close = Img('histogram_close.png')
    ws_close.add_image(img_close, 'J2')

    img_open = Img('histogram_open.png')
    ws_open.add_image(img_open, 'J2')

    wb.save('new_data.xlsx')
def sixth():
    df = pd.read_csv('preprocessed_data.csv', delimiter=',', header=None, names=['updated', 'high', 'low', 'open', 'close', 'volume', 'average', 'U/R'])

    df['volume'] = pd.to_numeric(df['volume'], errors='coerce')

    df.sort_values(by='volume', inplace=True)

    df['volume_decile'] = pd.qcut(df['volume'], 10, labels=False, duplicates='drop')

    min_first_decile = df[df['volume_decile'] == 0]['updated'].min()
    max_first_decile = df[df['volume_decile'] == 0]['updated'].max()
    min_tenth_decile = df[df['volume_decile'] == 9]['updated'].min()
    max_tenth_decile = df[df['volume_decile'] == 9]['updated'].max()

    results = pd.DataFrame({
        'Дециль': ['Первый', 'Десятый'],
        'Минимальная дата': [min_first_decile, min_tenth_decile],
        'Максимальная дата': [max_first_decile, max_tenth_decile]
    })

    wb = openpyxl.load_workbook('new_data.xlsx')

    ws = wb.create_sheet('Даты по децилям')

    for r in dataframe_to_rows(results, index=False, header=True):
        ws.append(r)

    wb.save('new_data.xlsx')

    decile_counts = df['volume_decile'].value_counts().sort_index()
    print("Количество данных в каждом дециле:")
    print(decile_counts)

def seventh():
    df = pd.read_csv('preprocessed_data.csv', delimiter=',', header=None,
                     names=['updated', 'high', 'low', 'open', 'close', 'volume', 'average', 'U/R'])

    # Словарь для перевода дней недели на английский
    days_translation = {
        'Понедельник': 'Monday',
        'Вторник': 'Tuesday',
        'Среда': 'Wednesday',
        'Четверг': 'Thursday',
        'Пятница': 'Friday',
        'Суббота': 'Saturday',
        'Воскресенье': 'Sunday'
    }

    # Замена русских названий дней недели на английские
    df['updated'] = df['updated'].replace(days_translation, regex=True)

    df['updated'] = pd.to_datetime(df['updated'], format='%A %d.%m.%Yг.')

    df['close'] = pd.to_numeric(df['close'], errors='coerce')
    df['open'] = pd.to_numeric(df['open'], errors='coerce')
    # Добавление столбцов с годом и месяцем
    df['year'] = df['updated'].dt.year
    df['month'] = df['updated'].dt.month

    # Сводная таблица по годам
    year_groups = df.groupby('year').size().reset_index(name='days_count')

    year_agg = df.groupby('year').apply(calculate_year_aggregates).reset_index()

    # Подписи к столбцам
    year_agg = year_agg.rename(columns={
        'year': 'Год',
        'close_higher_days': 'Количество дней, когда средняя цена превышала цену закрытия',
        'close_lower_days': 'Количество дней, когда средняя цена была меньше цены закрытия',
        'close_open_lower_days': 'Количество дней, когда цена закрытия была меньше цены открытия',
        'close_open_higher_days': 'Количество дней, когда цена закрытия была больше цены открытия'
    })

    # Сводная таблица по месяцам
    month_agg = df.groupby(['year', 'month']).apply(calculate_month_aggregates).reset_index()

    # Подпись к столбцу
    month_agg = month_agg.rename(columns={
        'close_higher_50_percent_days': 'Количество дней, когда средний объем торгов превышался на 50% по критерию U/R'
    })

    # Запись результатов в файл Excel
    with pd.ExcelWriter('new_data.xlsx', mode='a', engine='openpyxl') as writer:
        year_groups.to_excel(writer, sheet_name='Сводная таблица по годам1', index=False )
        year_agg.to_excel(writer, sheet_name='Сводная таблица по годам(агрегаты)', index=False, startcol=3)
        month_agg.to_excel(writer, sheet_name='Сводная таблица по месяцам1', index=False)


# Функция для агрегации по годам
def calculate_year_aggregates(group):
    return pd.Series({
        'close_higher_days': (group['close'] > group['close'].shift()).sum(),
        'close_lower_days': (group['close'] < group['close'].shift()).sum(),
        'close_open_lower_days': (
                    (group['close'] < group['close'].shift()) & (group['close'] < group['open'].shift())).sum(),
        'close_open_higher_days': (
                    (group['close'] > group['close'].shift()) & (group['close'] > group['open'].shift())).sum()
    })


# Функция для агрегации по месяцам
def calculate_month_aggregates(group):
    return pd.Series({
        'close_higher_50_percent_days': (group['volume'] > group['volume'].shift() * 1.5).sum()
    })


def eighth():
    # Чтение данных из CSV файла
    df = pd.read_csv('preprocessed_data.csv', delimiter=',', header=None,
                     names=['updated', 'high', 'low', 'open', 'close', 'volume', 'average', 'U/R'])

    # Словарь для перевода дней недели на английский
    days_translation = {
        'Понедельник': 'Monday',
        'Вторник': 'Tuesday',
        'Среда': 'Wednesday',
        'Четверг': 'Thursday',
        'Пятница': 'Friday',
        'Суббота': 'Saturday',
        'Воскресенье': 'Sunday'
    }

    # Замена русских названий дней недели на английские
    df['updated'] = df['updated'].replace(days_translation, regex=True)

    # Преобразование столбца 'updated' в формат даты
    df['updated'] = pd.to_datetime(df['updated'], format='%A %d.%m.%Yг.')

    # Преобразование строковых значений в числа
    numeric_columns = ['high', 'low', 'open', 'close', 'volume', 'average', 'U/R']
    df[numeric_columns] = df[numeric_columns].apply(pd.to_numeric, errors='coerce')

    # Удаление строк с неверными значениями
    df = df.dropna()

    # Создание столбцов для года и месяца
    df['year'] = df['updated'].dt.year
    df['month'] = df['updated'].dt.month

    monthly_aggregated = df.groupby(['year', 'month']).agg({
        'high': 'max',
        'low': 'min',
        'open': 'first',
        'close': 'last',
        'volume': 'sum',
        'average': 'mean',
        'U/R': 'sum'
    })

    with pd.ExcelWriter('new_data.xlsx', mode='a') as writer:
        monthly_aggregated.to_excel(writer, sheet_name='Сжатые(по месяцам)', index=True)
fifth()
sixth()
seventh()
eighth()






print("complete")