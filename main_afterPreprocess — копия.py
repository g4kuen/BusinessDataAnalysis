import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, Border, Alignment, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as Img

df = pd.read_csv('preprocessed_data.csv', delimiter=',', header=None,
                 names=['updated', 'high', 'low', 'open', 'close', 'volume', 'average', 'U/R'])
def fifth():
    df = pd.read_csv('preprocessed_data.csv', delimiter=',', header=None,
                     names=['updated', 'high', 'low', 'open', 'close', 'volume', 'average', 'U/R'])

    df['close'] = pd.to_numeric(df['close'], errors='coerce')
    df['open'] = pd.to_numeric(df['open'], errors='coerce')
    df_close_higher = df[df['close'] > df['open']]
    df_open_higher = df[df['open'] > df['close']]
    print(df[df['volume'] == '0'])
    wb = openpyxl.Workbook()


    header_font = Font(bold=True, italic=True)
    header_border = Border(bottom=Side(border_style='thin'))


    cell_alignment = Alignment(horizontal='center', vertical='center')
    cell_border = Border(top=Side(border_style='thin'),
                         bottom=Side(border_style='thin'),
                         left=Side(border_style='thin'),
                         right=Side(border_style='thin'))

    ws_close = wb.active
    ws_close.title = "Close higher"

    for r_idx, row in enumerate(dataframe_to_rows(df_close_higher, index=False), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_close.cell(row=r_idx, column=c_idx, value=value)
            cell.font = header_font
            cell.border = header_border
            cell.alignment = cell_alignment

    for row in ws_close.iter_rows(min_row=2, max_row=len(df_close_higher) + 1, min_col=1, max_col=len(df_close_higher.columns)):
        for cell in row:
            cell.border = cell_border
            cell.alignment = cell_alignment
    ws_close.freeze_panes = 'A2'
    ws_open = wb.create_sheet("Open higher")


    for r_idx, row in enumerate(dataframe_to_rows(df_open_higher, index=False), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_open.cell(row=r_idx, column=c_idx, value=value)
            cell.font = header_font
            cell.border = header_border
            cell.alignment = cell_alignment

    for row in ws_open.iter_rows(min_row=2, max_row=len(df_open_higher) + 1, min_col=1, max_col=len(df_open_higher.columns)):
        for cell in row:
            cell.border = cell_border
            cell.alignment = cell_alignment
    ws_close.freeze_panes = 'A2'
    fig_close, ax_close = plt.subplots()
    ax_close.hist(df_close_higher['volume'], bins=50)
    ax_close.set_xlabel('Объем торгов')
    ax_close.set_ylabel('Частота')
    ax_close.set_title('Гистограмма объемов торгов для дат, у которых точка закрытия выше точки открытия')
    plt.savefig('histogram_close.png')

    ws_close.freeze_panes = 'A2'
    fig_open, ax_open = plt.subplots()
    ax_open.hist(df_open_higher['volume'], bins=50)
    ax_open.set_xlabel('Объем торгов')
    ax_open.set_ylabel('Частота')
    ax_open.set_title('Гистограмма объемов торгов для дат, у которых точка закрытия ниже точки открытия')
    plt.savefig('histogram_open.png')

    img_close = Img('histogram_close.png')
    ws_close.add_image(img_close, 'J2')

    img_open = Img('histogram_open.png')
    ws_open.add_image(img_open, 'J2')
    ws_open.freeze_panes = 'A2'
    for row in ws_open.iter_rows(min_row=1, max_row=ws_open.max_row, min_col=1, max_col=ws_open.max_column):
        for cell in row:
            cell.border = cell_border
            cell.alignment = cell_alignment

    wb.save('new_data.xlsx')

    print("complete5")


def sixth():
    df = pd.read_csv('preprocessed_data.csv', delimiter=',', header=None, names=['updated', 'high', 'low', 'open', 'close', 'volume', 'average', 'U/R'])

    df['volume'] = pd.to_numeric(df['volume'], errors='coerce')

    df.sort_values(by='volume', inplace=True)

    df['volume_decile'] = pd.qcut(df['volume'], 10, labels=False, duplicates='drop')

    min_first_decile = df[df['volume_decile'] == 0]['updated'].min()
    max_first_decile = df[df['volume_decile'] == 0]['updated'].max()
    min_tenth_decile = df[df['volume_decile'] == 9]['updated'].min()
    max_tenth_decile = df[df['volume_decile'] == 9]['updated'].max()

    results = pd.DataFrame({
        'Дециль': ['Первый', 'Десятый'],
        'Минимальная дата': [min_first_decile, min_tenth_decile],
        'Максимальная дата': [max_first_decile, max_tenth_decile]
    })

    wb = openpyxl.load_workbook('new_data.xlsx')

    ws = wb.create_sheet('Даты по децилям')

    for r in dataframe_to_rows(results, index=False, header=True):
        ws.append(r)

    for row in ws.iter_rows(min_row=1, max_row=len(results) + 1, min_col=1, max_col=len(results.columns)):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            cell.border = Border(top=Side(border_style='thin'),
                                 bottom=Side(border_style='thin'),
                                 left=Side(border_style='thin'),
                                 right=Side(border_style='thin'))

    wb.save('new_data.xlsx')

    decile_counts = df['volume_decile'].value_counts().sort_index()
    print("Количество данных в каждом дециле:")
    print(decile_counts)

def seventh():
    df = pd.read_csv('preprocessed_data.csv', delimiter=',', header=None,
                     names=['updated', 'high', 'low', 'open', 'close', 'volume', 'average', 'U/R'])

    days_translation = {
        'Понедельник': 'Monday',
        'Вторник': 'Tuesday',
        'Среда': 'Wednesday',
        'Четверг': 'Thursday',
        'Пятница': 'Friday',
        'Суббота': 'Saturday',
        'Воскресенье': 'Sunday'
    }


    df['updated'] = df['updated'].replace(days_translation, regex=True)

    df.dropna(inplace=True)

    df['updated'] = pd.to_datetime(df['updated'], format='%A %d.%m.%Yг.')

    df['close'] = pd.to_numeric(df['close'], errors='coerce')
    df['open'] = pd.to_numeric(df['open'], errors='coerce')

    df['year'] = df['updated'].dt.year
    df['month'] = df['updated'].dt.month

    year_groups = df.groupby('year').size().reset_index(name='days_count')

    year_agg = df.groupby('year').apply(calculate_year_aggregates).reset_index()

    year_agg = year_agg.rename(columns={
        'year': 'Год',
        'close_higher_days': 'Количество дней, когда средняя цена превышала цену закрытия',
        'close_lower_days': 'Количество дней, когда средняя цена была меньше цены закрытия',
        'close_open_lower_days': 'Количество дней, когда цена закрытия была меньше цены открытия',
        'close_open_higher_days': 'Количество дней, когда цена закрытия была больше цены открытия'
    })

    month_agg = df.groupby(['year', 'month']).apply(calculate_month_aggregates).reset_index()

    month_agg = month_agg.rename(columns={
        'close_higher_50_percent_days': 'Количество дней, когда средний объем торгов превышался на 50% по критерию U/R'
    })

    wb = openpyxl.load_workbook('new_data.xlsx')

    ws_year_groups = wb.active
    ws_year_groups.title = "Сводная таблица по годам"

    for r in dataframe_to_rows(year_groups, index=False, header=True):
        ws_year_groups.append(r)

    for row in ws_year_groups.iter_rows(min_row=1, max_row=len(year_groups) + 1, min_col=1, max_col=len(year_groups.columns)):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            cell.border = Border(top=Side(border_style='thin'),
                                 bottom=Side(border_style='thin'),
                                 left=Side(border_style='thin'),
                                 right=Side(border_style='thin'))

    ws_year_groups.freeze_panes = 'A2'

    ws_year_agg = wb.create_sheet("Сводная таблица по годам (агрегаты)")

    for r in dataframe_to_rows(year_agg, index=False, header=True):
        ws_year_agg.append(r)

    for row in ws_year_agg.iter_rows(min_row=1, max_row=len(year_agg) + 1, min_col=1, max_col=len(year_agg.columns)):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            cell.border = Border(top=Side(border_style='thin'),
                                 bottom=Side(border_style='thin'),
                                 left=Side(border_style='thin'),
                                 right=Side(border_style='thin'))

    ws_year_agg.freeze_panes = 'A2'

    ws_month_agg = wb.create_sheet("Сводная таблица по месяцам")

    for r in dataframe_to_rows(month_agg, index=False, header=True):
        ws_month_agg.append(r)

    for row in ws_month_agg.iter_rows(min_row=1, max_row=len(month_agg) + 1, min_col=1, max_col=len(month_agg.columns)):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            cell.border = Border(top=Side(border_style='thin'),
                                 bottom=Side(border_style='thin'),
                                 left=Side(border_style='thin'),
                                 right=Side(border_style='thin'))

    ws_month_agg.freeze_panes = 'A2'

    wb.save('new_data.xlsx')

    print("complete7")


def calculate_year_aggregates(group):
    return pd.Series({
        'close_higher_days': (group['close'] > group['close'].shift()).sum(),
        'close_lower_days': (group['close'] < group['close'].shift()).sum(),
        'close_open_lower_days': (
                    (group['close'] < group['close'].shift()) & (group['close'] < group['open'].shift())).sum(),
        'close_open_higher_days': (
                    (group['close'] > group['close'].shift()) & (group['close'] > group['open'].shift())).sum()
    })


def calculate_month_aggregates(group):
    return pd.Series({
        'close_higher_50_percent_days': (group['volume'] > group['volume'].shift() * 1.5).sum()
    })

def format_sheet(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True, italic=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(top=Side(border_style='medium'),
                                 bottom=Side(border_style='medium'),
                                 left=Side(border_style='medium'),
                                 right=Side(border_style='medium'))

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(top=Side(border_style='thin'),
                                 bottom=Side(border_style='thin'),
                                 left=Side(border_style='thin'),
                                 right=Side(border_style='thin'))


def eighth():
    df = pd.read_csv('preprocessed_data.csv', delimiter=',', header=None,
                     names=['updated', 'high', 'low', 'open', 'close', 'volume', 'average', 'U/R'])

    days_translation = {
        'Понедельник': 'Monday',
        'Вторник': 'Tuesday',
        'Среда': 'Wednesday',
        'Четверг': 'Thursday',
        'Пятница': 'Friday',
        'Суббота': 'Saturday',
        'Воскресенье': 'Sunday'
    }

    df['updated'] = df['updated'].replace(days_translation, regex=True)

    df['updated'] = pd.to_datetime(df['updated'], format='%A %d.%m.%Yг.')

    numeric_columns = ['high', 'low', 'open', 'close', 'volume', 'average', 'U/R']
    df[numeric_columns] = df[numeric_columns].apply(pd.to_numeric, errors='coerce')

    df = df.dropna()

    df['year'] = df['updated'].dt.year
    df['month'] = df['updated'].dt.month

    monthly_aggregated = df.groupby(['year', 'month']).agg({
        'high': 'max',
        'low': 'min',
        'open': 'first',
        'close': 'last',
        'volume': 'sum',
        'average': 'mean',
        'U/R': 'sum'
    })

    with pd.ExcelWriter('new_data.xlsx', mode='a', engine='openpyxl') as writer:
        monthly_aggregated.to_excel(writer, sheet_name='Сжатые (по месяцам)', index=True)

    wb = openpyxl.load_workbook('new_data.xlsx')

    ws_monthly_aggregated = wb['Сжатые (по месяцам)']


    format_sheet(ws_monthly_aggregated)


    ws_monthly_aggregated.freeze_panes = 'A2'

    wb.save('new_data.xlsx')

    print("complete8")



df = pd.read_csv('preprocessed_data.csv', skiprows=[0])
df.to_csv('preprocessed_data.csv', index=False)
fifth()
sixth()
seventh()
eighth()






print("complete")